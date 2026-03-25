"""Render memo data to an Excel workbook matching the 不動産モデルTemplate.en.xlsx format.

Five sheets:
  1. Cover             — title page
  2. Input             — assumptions & parameters
  3. Property Summary  — asset summary, metrics, DCF table, sensitivity
  4. CFs               — detailed annual cash flow model
  5. Sources & Uses    — capital structure breakdown
"""
from __future__ import annotations

from datetime import date
from typing import Optional

import numpy_financial as npf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------
TEAL_DARK = "FF005B59"   # section header background
TEAL_MID = "FFD9E8E6"    # alternate rows / metric rows
TEAL_LIGHT = "FFE3F2EF"  # input cell background
GRAY_CALC = "FFCCCDCB"   # calculated value background
YELLOW_KEY = "FFFFFF00"  # key figure highlight
WHITE = "FFFFFFFF"
BLACK = "FF000000"

# ---------------------------------------------------------------------------
# Number formats
# ---------------------------------------------------------------------------
FMT_CURRENCY = "$#,##0"
FMT_PERCENT = "0.00%"
FMT_MULTIPLE = '0.00"x"'
FMT_INT = "#,##0"

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _font(bold: bool = False, white: bool = False, size: int = 10) -> Font:
    color = WHITE if white else BLACK
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _align(horizontal: str = "left", vertical: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


def _thin_border() -> Border:
    thin = Side(style="thin", color="FFB0B0B0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_section_header(ws, row: int, col: int, text: str, span: int = 2) -> None:
    """Write a TEAL_DARK section header, optionally merged across span columns."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = _fill(TEAL_DARK)
    cell.font = _font(bold=True, white=True, size=10)
    cell.alignment = _align("left")
    if span > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + span - 1
        )


def _write_label_value(
    ws,
    row: int,
    label_col: int,
    label: str,
    value,
    value_col: int,
    bold_label: bool = False,
    bold_value: bool = False,
    value_fill: str = TEAL_LIGHT,
    number_format: Optional[str] = None,
) -> None:
    """Write a label / value pair with appropriate styling."""
    lc = ws.cell(row=row, column=label_col, value=label)
    lc.font = _font(bold=bold_label)
    lc.alignment = _align("left")

    vc = ws.cell(row=row, column=value_col, value=value)
    vc.font = _font(bold=bold_value)
    vc.fill = _fill(value_fill)
    vc.alignment = _align("right")
    if number_format:
        vc.number_format = number_format


def _set_tab_color(ws, color: str = "005B59") -> None:
    ws.sheet_properties.tabColor = color


def _set_col_widths(ws, widths: dict) -> None:
    """widths: {col_letter_or_int: width}"""
    for col, width in widths.items():
        letter = col if isinstance(col, str) else get_column_letter(col)
        ws.column_dimensions[letter].width = width


# ---------------------------------------------------------------------------
# Calculator imports (lazy-safe inline reimport)
# ---------------------------------------------------------------------------

def _calc_imports():
    from memo_generator.financials.calculator import (
        calculate_annual_debt_service,
        calculate_going_in_cap_rate,
        calculate_noi,
        _remaining_loan_balance,
    )
    return (
        calculate_noi,
        calculate_going_in_cap_rate,
        calculate_annual_debt_service,
        _remaining_loan_balance,
    )


def _safe_irr(cash_flows: list[float]) -> Optional[float]:
    try:
        result = npf.irr(cash_flows)
        if result is None or result != result:
            return None
        return float(result)
    except Exception:
        return None


def _levered_irr_override(
    prop,
    exit_cap_override: Optional[float] = None,
    purchase_price_override: Optional[float] = None,
    hold_override: Optional[int] = None,
) -> Optional[float]:
    """Recompute levered IRR with selective parameter overrides."""
    calc_noi, calc_cap, calc_ds, calc_rem = _calc_imports()

    purchase_price = purchase_price_override if purchase_price_override is not None else prop.purchase_price
    hold = hold_override if hold_override is not None else prop.hold_period_years

    # Equity is scaled proportionally when purchase price changes
    equity = purchase_price - prop.loan_amount

    noi_yr0 = calc_noi(prop.gross_rental_income, prop.vacancy_rate, prop.operating_expenses)
    annual_ds = calc_ds(prop.loan_amount, prop.interest_rate, prop.amortization_years)
    base_cap = calc_cap(noi_yr0, prop.purchase_price)
    exit_cap = exit_cap_override if exit_cap_override is not None else (
        prop.exit_cap_rate if prop.exit_cap_rate is not None else base_cap
    )

    if equity <= 0 or exit_cap <= 0:
        return None

    cash_flows: list[float] = [-equity]
    for year in range(1, hold + 1):
        noi = noi_yr0 * ((1 + prop.rent_growth_annual) ** year)
        capex = prop.capital_expenditures * ((1 + prop.expense_growth_annual) ** year)
        cf = noi - annual_ds - capex
        cash_flows.append(cf)

    noi_exit = noi_yr0 * ((1 + prop.rent_growth_annual) ** hold)
    exit_value = noi_exit / exit_cap
    remaining_debt = calc_rem(prop.loan_amount, prop.interest_rate, prop.amortization_years, hold)
    cash_flows[-1] += exit_value - remaining_debt

    return _safe_irr(cash_flows)


# ---------------------------------------------------------------------------
# Sheet 1: Cover
# ---------------------------------------------------------------------------

def _build_cover(wb: Workbook, prop) -> None:
    ws = wb.create_sheet("Cover")
    _set_tab_color(ws)

    ws.sheet_view.showGridLines = False

    # Column widths
    _set_col_widths(ws, {"A": 4, "B": 20, "C": 20, "D": 20, "E": 20, "F": 20, "G": 30, "H": 20})

    # Row heights
    for r in range(1, 20):
        ws.row_dimensions[r].height = 20

    # G6: "Investment Analysis"
    c6 = ws.cell(row=6, column=7, value="Investment Analysis")
    c6.font = Font(bold=True, size=22, color="FF005B59", name="Calibri")
    c6.alignment = _align("center")

    # G12: property name
    c12 = ws.cell(row=12, column=7, value=prop.property_name)
    c12.font = Font(bold=True, size=18, color=BLACK, name="Calibri")
    c12.alignment = _align("center")

    # G14: city, state
    c14 = ws.cell(row=14, column=7, value=f"{prop.city}, {prop.state_or_country}")
    c14.font = Font(size=14, color="FF444444", name="Calibri")
    c14.alignment = _align("center")

    # G16: date
    c16 = ws.cell(row=16, column=7, value=date.today().strftime("%B %Y"))
    c16.font = Font(size=12, color="FF777777", name="Calibri")
    c16.alignment = _align("center")


# ---------------------------------------------------------------------------
# Sheet 2: Input
# ---------------------------------------------------------------------------

def _build_input(wb: Workbook, prop) -> None:
    ws = wb.create_sheet("Input")
    _set_tab_color(ws)
    ws.sheet_view.showGridLines = False

    _set_col_widths(ws, {"A": 4, "B": 32, "C": 16, "D": 14})

    B, C = 2, 3  # label col, value col

    # Row 1: sheet title
    _write_section_header(ws, 1, B, "Input — Assumptions & Parameters", span=2)

    # Row 3: Property Information
    _write_section_header(ws, 3, B, "Property Information", span=2)
    _write_label_value(ws, 4, B, "Project Name", prop.property_name, C)
    _write_label_value(ws, 5, B, "Location",
                       f"{prop.address}, {prop.city}, {prop.state_or_country}", C)
    _write_label_value(ws, 6, B, "Usage", prop.property_type, C)
    _write_label_value(ws, 7, B, "Year Built", prop.year_built, C,
                       number_format=FMT_INT)
    _write_label_value(ws, 8, B, "Total Units",
                       prop.total_units if prop.total_units else "N/A", C,
                       number_format=FMT_INT if prop.total_units else None)
    _write_label_value(ws, 9, B, "Total GFA (sqft)",
                       prop.total_sqft if prop.total_sqft else "N/A", C,
                       number_format=FMT_INT if prop.total_sqft else None)

    # Row 11: Acquisition Terms
    _write_section_header(ws, 11, B, "Acquisition Terms", span=2)
    _write_label_value(ws, 12, B, "Purchase Price", prop.purchase_price, C,
                       bold_value=True, value_fill=YELLOW_KEY, number_format=FMT_CURRENCY)
    _write_label_value(ws, 13, B, "Loan Amount", prop.loan_amount, C,
                       number_format=FMT_CURRENCY)
    _write_label_value(ws, 14, B, "Equity Invested", prop.equity_invested, C,
                       number_format=FMT_CURRENCY)

    # Row 16: Assumptions
    _write_section_header(ws, 16, B, "Assumptions", span=2)
    _write_label_value(ws, 17, B, "Holding Period",
                       f"{prop.hold_period_years} years", C, bold_value=True,
                       value_fill=YELLOW_KEY)
    _write_label_value(ws, 18, B, "Vacancy Rate", prop.vacancy_rate, C,
                       number_format=FMT_PERCENT)
    _write_label_value(ws, 19, B, "Rent Growth (annual)", prop.rent_growth_annual, C,
                       number_format=FMT_PERCENT)
    _write_label_value(ws, 20, B, "Expense Growth (annual)", prop.expense_growth_annual, C,
                       number_format=FMT_PERCENT)
    _write_label_value(ws, 21, B, "Exit Cap Rate",
                       prop.exit_cap_rate if prop.exit_cap_rate is not None else "N/A", C,
                       bold_value=True, value_fill=YELLOW_KEY,
                       number_format=FMT_PERCENT if prop.exit_cap_rate is not None else None)
    _write_label_value(ws, 22, B, "Discount Rate", "N/A", C)

    # Row 24: Financing
    _write_section_header(ws, 24, B, "Financing", span=2)
    _write_label_value(ws, 25, B, "Interest Rate", prop.interest_rate, C,
                       number_format=FMT_PERCENT)
    _write_label_value(ws, 26, B, "Loan Term", f"{prop.loan_term_years} years", C)
    _write_label_value(ws, 27, B, "Amortization", f"{prop.amortization_years} years", C)

    # Row 29: Operating
    _write_section_header(ws, 29, B, "Operating", span=2)
    _write_label_value(ws, 30, B, "Gross Rental Income (annual)", prop.gross_rental_income, C,
                       number_format=FMT_CURRENCY)
    _write_label_value(ws, 31, B, "Operating Expenses (annual)", prop.operating_expenses, C,
                       number_format=FMT_CURRENCY)
    _write_label_value(ws, 32, B, "CapEx Reserve (annual)", prop.capital_expenditures, C,
                       number_format=FMT_CURRENCY)


# ---------------------------------------------------------------------------
# Sheet 3: Property Summary
# ---------------------------------------------------------------------------

def _build_property_summary(wb: Workbook, prop, metrics: dict) -> None:  # noqa: C901
    calc_noi, calc_cap, calc_ds, calc_rem = _calc_imports()

    ws = wb.create_sheet("Property Summary")
    _set_tab_color(ws)
    ws.sheet_view.showGridLines = False

    hold = prop.hold_period_years
    noi_yr0 = calc_noi(prop.gross_rental_income, prop.vacancy_rate, prop.operating_expenses)
    annual_ds = calc_ds(prop.loan_amount, prop.interest_rate, prop.amortization_years)
    base_cap = calc_cap(noi_yr0, prop.purchase_price)
    exit_cap = prop.exit_cap_rate if prop.exit_cap_rate is not None else base_cap

    # Column widths: A=4, B=32, C=16, D=14, E=14, F=32, G=16, H=14, I=14, J=32, K=16
    _set_col_widths(ws, {
        "A": 4, "B": 32, "C": 16, "D": 14, "E": 14,
        "F": 32, "G": 16, "H": 14, "I": 14,
        "J": 32, "K": 16,
    })
    # Year columns starting at D (col 4) for DCF below
    for col_idx in range(4, 4 + hold + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    B, C = 2, 3
    F, G = 6, 7
    J, K = 10, 11

    # ---- Row 1: sheet title spanning full width ----
    _write_section_header(ws, 1, B, f"Property Summary — {prop.property_name}", span=10)

    # ======================================================
    # LEFT BLOCK (B-E): Asset Summary
    # ======================================================
    _write_section_header(ws, 3, B, "Asset Summary", span=4)

    left_asset = [
        (4,  "Location",                     prop.address,                          False, TEAL_LIGHT, None),
        (5,  "City",                          f"{prop.city}, {prop.state_or_country}", False, TEAL_LIGHT, None),
        (6,  "Usage",                         prop.property_type,                    False, TEAL_LIGHT, None),
        (7,  "Year Built",                    prop.year_built,                       False, TEAL_LIGHT, FMT_INT),
        (8,  "Total Units",                   prop.total_units or "N/A",             False, TEAL_LIGHT,
             FMT_INT if prop.total_units else None),
        (9,  "Total GFA (sqft)",              prop.total_sqft or "N/A",              False, TEAL_LIGHT,
             FMT_INT if prop.total_sqft else None),
    ]
    for row, label, value, bold, vfill, fmt in left_asset:
        _write_label_value(ws, row, B, label, value, C, bold_value=bold, value_fill=vfill, number_format=fmt)

    _write_section_header(ws, 11, B, "Acquisition Info (USD)", span=4)

    ltv_val = metrics.get("ltv", 0.0)
    left_acq = [
        (12, "Purchase Price (excl. tax)", prop.purchase_price,    True,  YELLOW_KEY, FMT_CURRENCY),
        (13, "Loan Amount",                prop.loan_amount,        False, TEAL_LIGHT, FMT_CURRENCY),
        (14, "Required Equity",            prop.equity_invested,    True,  YELLOW_KEY, FMT_CURRENCY),
        (15, "LTV",                        ltv_val,                 True,  GRAY_CALC,  FMT_PERCENT),
        (16, "Loan Interest",              prop.interest_rate,      False, TEAL_LIGHT, FMT_PERCENT),
        (17, "Loan Term",                  f"{prop.loan_term_years} years", False, TEAL_LIGHT, None),
    ]
    for row, label, value, bold, vfill, fmt in left_acq:
        _write_label_value(ws, row, B, label, value, C, bold_value=bold, value_fill=vfill, number_format=fmt)

    # ======================================================
    # CENTRE BLOCK (F-I): Assumptions + Sources & Uses
    # ======================================================
    _write_section_header(ws, 3, F, "Assumptions", span=4)

    egi = prop.gross_rental_income * (1 - prop.vacancy_rate)
    opex_ratio = prop.operating_expenses / egi if egi else 0.0
    capex_ratio = prop.capital_expenditures / noi_yr0 if noi_yr0 else 0.0

    centre_assump = [
        (4,  "Holding Period",          prop.hold_period_years,          True,  YELLOW_KEY, FMT_INT),
        (5,  "Vacancy Rate",            prop.vacancy_rate,               False, TEAL_LIGHT, FMT_PERCENT),
        (6,  "Rent Growth (annual)",    prop.rent_growth_annual,          False, TEAL_LIGHT, FMT_PERCENT),
        (7,  "Exit Cap Rate",           exit_cap,                        True,  YELLOW_KEY, FMT_PERCENT),
        (8,  "OPEX Ratio (vs EGI)",     opex_ratio,                      False, GRAY_CALC,  FMT_PERCENT),
        (9,  "CapEx Ratio (vs NOI)",    capex_ratio,                     False, GRAY_CALC,  FMT_PERCENT),
    ]
    for row, label, value, bold, vfill, fmt in centre_assump:
        _write_label_value(ws, row, F, label, value, G, bold_value=bold, value_fill=vfill, number_format=fmt)

    _write_section_header(ws, 11, F, "Sources & Uses", span=4)

    acq_costs_est = prop.purchase_price * 0.05
    total_uses = prop.purchase_price + acq_costs_est
    total_sources = prop.loan_amount + prop.equity_invested
    diff = total_uses - total_sources

    su_rows = [
        (12, "Uses",                    "Total Uses",                False, TEAL_MID,   None),
        (13, "  Acquisition Price",     prop.purchase_price,         False, TEAL_LIGHT, FMT_CURRENCY),
        (14, "  Acquisition Costs (est. 5%)", acq_costs_est,         False, TEAL_LIGHT, FMT_CURRENCY),
        (15, "Total Uses",              total_uses,                  True,  GRAY_CALC,  FMT_CURRENCY),
        (16, "",                        None,                        False, WHITE,      None),
        (17, "Sources",                 "",                          False, TEAL_MID,   None),
        (18, "  Senior Loan",           prop.loan_amount,            False, TEAL_LIGHT, FMT_CURRENCY),
        (19, "  Equity",                prop.equity_invested,        False, TEAL_LIGHT, FMT_CURRENCY),
        (20, "Total Sources",           total_sources,               True,  GRAY_CALC,  FMT_CURRENCY),
        (21, "Check (Diff)",            diff,                        False, GRAY_CALC,  FMT_CURRENCY),
    ]
    for row, label, value, bold, vfill, fmt in su_rows:
        lc = ws.cell(row=row, column=F, value=label)
        lc.font = _font(bold=bold)
        if value is not None:
            vc = ws.cell(row=row, column=G, value=value)
            vc.font = _font(bold=bold)
            vc.fill = _fill(vfill)
            vc.alignment = _align("right")
            if fmt:
                vc.number_format = fmt

    # ======================================================
    # RIGHT BLOCK (J-K): Investment Metrics Summary
    # ======================================================
    _write_section_header(ws, 3, J, "Investment Metrics Summary", span=2)

    gross_yield = prop.gross_rental_income / prop.purchase_price if prop.purchase_price else 0.0
    irr_val = metrics.get("irr")
    em_val = metrics.get("equity_multiple", 0.0)
    coc_val = metrics.get("cash_on_cash_yr1", 0.0)
    dscr_val = metrics.get("dscr", 0.0)

    # Unlevered IRR: CF = [-purchase_price, noi_yr1, ..., noi_yrN + exit]
    unlev_cfs: list[float] = [-prop.purchase_price]
    for yr in range(1, hold + 1):
        unlev_cfs.append(noi_yr0 * ((1 + prop.rent_growth_annual) ** yr))
    noi_exit = noi_yr0 * ((1 + prop.rent_growth_annual) ** hold)
    exit_val = noi_exit / exit_cap
    unlev_cfs[-1] += exit_val
    unlevered_irr = _safe_irr(unlev_cfs)

    metrics_right = [
        (4,  "Gross Yield",      gross_yield,    FMT_PERCENT),
        (5,  "NOI Yield",        base_cap,       FMT_PERCENT),
        (6,  "CoC Return",       coc_val,        FMT_PERCENT),
        (7,  "DSCR",             dscr_val,       '0.00"x"'),
        (8,  "Equity Multiple",  em_val,         FMT_MULTIPLE),
        (9,  "Unlevered IRR",    unlevered_irr,  FMT_PERCENT),
        (10, "Levered IRR",      irr_val,        FMT_PERCENT),
        (11, "Total Units",      prop.total_units or "N/A", FMT_INT if prop.total_units else None),
        (12, "Annual Rental Income", prop.gross_rental_income, FMT_CURRENCY),
    ]
    for row, label, value, fmt in metrics_right:
        lc = ws.cell(row=row, column=J, value=label)
        lc.font = _font()
        vc = ws.cell(row=row, column=K, value=value)
        vc.fill = _fill(TEAL_MID)
        vc.font = _font(bold=True)
        vc.alignment = _align("right")
        if fmt and value is not None and not isinstance(value, str):
            vc.number_format = fmt

    # ======================================================
    # LOWER SECTION: Simplified DCF (Annual, Levered)
    # ======================================================
    DCF_START = 40  # first row of DCF table

    dcf_title = ws.cell(row=DCF_START, column=B, value="Simplified DCF (Annual, Levered)")
    dcf_title.font = _font(bold=True, size=11)

    # Header row
    hdr_row = DCF_START + 1
    ws.cell(row=hdr_row, column=B, value="").fill = _fill(TEAL_DARK)
    ws.cell(row=hdr_row, column=C, value="Assumptions").fill = _fill(TEAL_DARK)
    ws.cell(row=hdr_row, column=C).font = _font(bold=True, white=True)

    col_year0 = 4  # column D = Year 0
    for yr in range(0, hold + 1):
        col = col_year0 + yr
        c = ws.cell(row=hdr_row, column=col, value=f"Year {yr}")
        c.fill = _fill(TEAL_DARK)
        c.font = _font(bold=True, white=True)
        c.alignment = _align("center")

    # Build per-year arrays
    gpi_vals = [prop.gross_rental_income]  # Year 0 assumption
    vac_rate_vals = [prop.vacancy_rate]
    vac_loss_vals = [0.0]
    egi_vals = [0.0]
    opex_vals = [prop.operating_expenses]
    noi_vals_dcf = [0.0]
    capex_vals = [prop.capital_expenditures]
    ncf_vals = [0.0]
    ds_vals = [0.0]
    lev_cf_vals = [0.0]

    for yr in range(1, hold + 1):
        gpi = prop.gross_rental_income * ((1 + prop.rent_growth_annual) ** yr)
        vac_loss = gpi * prop.vacancy_rate
        egi = gpi - vac_loss
        opex = prop.operating_expenses * ((1 + prop.expense_growth_annual) ** yr)
        noi = egi - opex
        capex = prop.capital_expenditures * ((1 + prop.expense_growth_annual) ** yr)
        ncf = noi - capex
        lev_cf = ncf - annual_ds

        gpi_vals.append(gpi)
        vac_rate_vals.append(prop.vacancy_rate)
        vac_loss_vals.append(-vac_loss)
        egi_vals.append(egi)
        opex_vals.append(-opex)
        noi_vals_dcf.append(noi)
        capex_vals.append(-capex)
        ncf_vals.append(ncf)
        ds_vals.append(-annual_ds)
        lev_cf_vals.append(lev_cf)

    # Exit values
    noi_exit_dcf = noi_yr0 * ((1 + prop.rent_growth_annual) ** hold)
    exit_value_dcf = noi_exit_dcf / exit_cap
    disp_costs = exit_value_dcf * 0.02
    net_sale_proceeds = exit_value_dcf - disp_costs
    remaining_debt = calc_rem(prop.loan_amount, prop.interest_rate, prop.amortization_years, hold)
    net_equity_proceeds = net_sale_proceeds - remaining_debt

    # Total CF to equity
    total_cf = [-prop.equity_invested]
    for yr in range(1, hold + 1):
        if yr < hold:
            total_cf.append(lev_cf_vals[yr])
        else:
            total_cf.append(lev_cf_vals[yr] + net_equity_proceeds)

    def _write_dcf_row(row: int, label: str, assumption_val, year_vals: list,
                       bold: bool = False, fmt: str = FMT_CURRENCY,
                       label_indent: bool = False) -> None:
        lbl = ("  " if label_indent else "") + label
        lc = ws.cell(row=row, column=B, value=lbl)
        lc.font = _font(bold=bold)

        ac = ws.cell(row=row, column=C, value=assumption_val)
        ac.font = _font(bold=bold)
        if assumption_val is not None and isinstance(assumption_val, (int, float)):
            ac.number_format = fmt
        ac.fill = _fill(GRAY_CALC)
        ac.alignment = _align("right")

        for yr_idx, val in enumerate(year_vals):
            col = col_year0 + yr_idx
            vc = ws.cell(row=row, column=col, value=val)
            vc.font = _font(bold=bold)
            vc.alignment = _align("right")
            if val is not None and isinstance(val, (int, float)):
                vc.number_format = fmt
            if bold:
                vc.fill = _fill(GRAY_CALC)

    r = DCF_START + 2
    _write_dcf_row(r, "GPI (Gross Potential Income)", gpi_vals[0], gpi_vals); r += 1
    _write_dcf_row(r, "  Vacancy Rate", vac_rate_vals[0], vac_rate_vals,
                   fmt=FMT_PERCENT, label_indent=False); r += 1
    _write_dcf_row(r, "  Vacancy Loss", None, vac_loss_vals, label_indent=False); r += 1
    _write_dcf_row(r, "EGI (Effective Gross Income)", None, egi_vals, bold=True); r += 1
    _write_dcf_row(r, "  OPEX (Operating Expenses)", opex_vals[0], opex_vals); r += 1
    _write_dcf_row(r, "NOI", None, noi_vals_dcf, bold=True); r += 1
    _write_dcf_row(r, "  CapEx", capex_vals[0], capex_vals); r += 1
    _write_dcf_row(r, "NCF", None, ncf_vals, bold=True); r += 1
    _write_dcf_row(r, "  Debt Service", -annual_ds, ds_vals); r += 1
    _write_dcf_row(r, "Levered Cash Flow", None, lev_cf_vals, bold=True); r += 1

    r += 1  # blank
    exit_row_vals = [0.0] * (hold + 1)
    exit_row_vals[hold] = exit_value_dcf
    disp_row_vals = [0.0] * (hold + 1)
    disp_row_vals[hold] = -disp_costs
    nsp_row_vals = [0.0] * (hold + 1)
    nsp_row_vals[hold] = net_sale_proceeds
    debt_rep_vals = [0.0] * (hold + 1)
    debt_rep_vals[hold] = -remaining_debt
    nep_row_vals = [0.0] * (hold + 1)
    nep_row_vals[hold] = net_equity_proceeds

    _write_dcf_row(r, "Sale Price (Exit Cap)", None, exit_row_vals); r += 1
    _write_dcf_row(r, "  Disposition Costs", None, disp_row_vals); r += 1
    _write_dcf_row(r, "Net Sale Proceeds", None, nsp_row_vals); r += 1
    _write_dcf_row(r, "  Loan Balance Repayment", None, debt_rep_vals); r += 1
    _write_dcf_row(r, "Net Equity Proceeds", None, nep_row_vals); r += 1

    r += 1  # blank
    _write_dcf_row(r, "Total CF (to Equity)", None, total_cf, bold=True)
    # Color Year 0 (equity outlay) yellow
    ws.cell(row=r, column=col_year0).fill = _fill(YELLOW_KEY)
    r += 1

    r += 1
    irr_row = r
    irr_lbl = ws.cell(row=irr_row, column=B, value="IRR (Levered)")
    irr_lbl.font = _font(bold=True)
    irr_cell = ws.cell(row=irr_row, column=C, value=irr_val)
    irr_cell.font = _font(bold=True)
    irr_cell.fill = _fill(YELLOW_KEY)
    irr_cell.alignment = _align("right")
    if irr_val is not None:
        irr_cell.number_format = FMT_PERCENT
    r += 1

    em_lbl = ws.cell(row=r, column=B, value="Equity Multiple")
    em_lbl.font = _font(bold=True)
    em_cell = ws.cell(row=r, column=C, value=em_val)
    em_cell.font = _font(bold=True)
    em_cell.fill = _fill(YELLOW_KEY)
    em_cell.alignment = _align("right")
    em_cell.number_format = FMT_MULTIPLE
    r += 2

    # ======================================================
    # RIGHT LOWER: Sensitivity Analysis (cols J+)
    # ======================================================
    SENS_START_ROW = DCF_START - 23  # row 17
    SENS_COL_LABEL = 10   # J
    SENS_COL_START = 11   # K .. O  (5 columns for 5 exit cap deltas)

    exit_cap_deltas = [-0.01, -0.005, 0.0, 0.005, 0.01]
    price_deltas = [-0.10, -0.05, 0.0, 0.05, 0.10]
    hold_deltas = [-2, -1, 0, 1, 2]

    sr = SENS_START_ROW

    # Table 1 header
    t1_title = ws.cell(row=sr, column=SENS_COL_LABEL,
                       value="Sensitivity Analysis")
    t1_title.font = _font(bold=True, size=11)
    sr += 1

    t1_sub = ws.cell(row=sr, column=SENS_COL_LABEL,
                     value="Purchase Price (×) × Exit Cap Rate → IRR")
    t1_sub.font = _font(bold=True)
    sr += 1

    # Column headers (exit cap deltas)
    ws.cell(row=sr, column=SENS_COL_LABEL, value="Price \\ Exit Cap").fill = _fill(TEAL_DARK)
    ws.cell(row=sr, column=SENS_COL_LABEL).font = _font(bold=True, white=True)
    ws.cell(row=sr, column=SENS_COL_LABEL).alignment = _align("center")
    for ci, delta in enumerate(exit_cap_deltas):
        col = SENS_COL_START + ci
        sign = "+" if delta >= 0 else ""
        hc = ws.cell(row=sr, column=col, value=f"{sign}{delta:.1%}")
        hc.fill = _fill(TEAL_DARK)
        hc.font = _font(bold=True, white=True)
        hc.alignment = _align("center")
    sr += 1

    for pi, pd in enumerate(price_deltas):
        test_price = prop.purchase_price * (1 + pd)
        sign = "+" if pd >= 0 else ""
        row_lbl = ws.cell(row=sr, column=SENS_COL_LABEL, value=f"{sign}{pd:.0%}")
        row_lbl.fill = _fill(TEAL_MID)
        row_lbl.font = _font(bold=True)
        row_lbl.alignment = _align("center")
        for ci, cd in enumerate(exit_cap_deltas):
            test_exit = exit_cap + cd
            if test_exit <= 0:
                ws.cell(row=sr, column=SENS_COL_START + ci, value="N/A")
                continue
            sens_irr = _levered_irr_override(prop,
                                             exit_cap_override=test_exit,
                                             purchase_price_override=test_price)
            vc = ws.cell(row=sr, column=SENS_COL_START + ci,
                         value=sens_irr if sens_irr is not None else "N/A")
            if sens_irr is not None:
                vc.number_format = FMT_PERCENT
            # highlight diagonal (base case)
            if pi == 2 and ci == 2:
                vc.fill = _fill(YELLOW_KEY)
            else:
                vc.fill = _fill(TEAL_LIGHT)
            vc.alignment = _align("center")
        sr += 1

    sr += 2  # spacer

    # Table 2: Holding Period × Exit Cap Rate
    t2_sub = ws.cell(row=sr, column=SENS_COL_LABEL,
                     value="Holding Period × Exit Cap Rate → Levered IRR")
    t2_sub.font = _font(bold=True)
    sr += 1

    ws.cell(row=sr, column=SENS_COL_LABEL, value="Hold \\ Exit Cap").fill = _fill(TEAL_DARK)
    ws.cell(row=sr, column=SENS_COL_LABEL).font = _font(bold=True, white=True)
    ws.cell(row=sr, column=SENS_COL_LABEL).alignment = _align("center")
    for ci, delta in enumerate(exit_cap_deltas):
        col = SENS_COL_START + ci
        sign = "+" if delta >= 0 else ""
        hc = ws.cell(row=sr, column=col, value=f"{sign}{delta:.1%}")
        hc.fill = _fill(TEAL_DARK)
        hc.font = _font(bold=True, white=True)
        hc.alignment = _align("center")
    sr += 1

    for hi, hd in enumerate(hold_deltas):
        test_hold = max(1, hold + hd)
        sign = "+" if hd >= 0 else ""
        row_lbl = ws.cell(row=sr, column=SENS_COL_LABEL, value=f"{sign}{hd}yr ({test_hold}yr)")
        row_lbl.fill = _fill(TEAL_MID)
        row_lbl.font = _font(bold=True)
        row_lbl.alignment = _align("center")
        for ci, cd in enumerate(exit_cap_deltas):
            test_exit = exit_cap + cd
            if test_exit <= 0:
                ws.cell(row=sr, column=SENS_COL_START + ci, value="N/A")
                continue
            sens_irr = _levered_irr_override(prop,
                                             exit_cap_override=test_exit,
                                             hold_override=test_hold)
            vc = ws.cell(row=sr, column=SENS_COL_START + ci,
                         value=sens_irr if sens_irr is not None else "N/A")
            if sens_irr is not None:
                vc.number_format = FMT_PERCENT
            if hi == 2 and ci == 2:
                vc.fill = _fill(YELLOW_KEY)
            else:
                vc.fill = _fill(TEAL_LIGHT)
            vc.alignment = _align("center")
        sr += 1


# ---------------------------------------------------------------------------
# Sheet 4: CFs (Detailed Annual Cash Flow Model)
# ---------------------------------------------------------------------------

def _build_cfs(wb: Workbook, prop, metrics: dict) -> None:
    calc_noi, calc_cap, calc_ds, calc_rem = _calc_imports()

    ws = wb.create_sheet("CFs")
    _set_tab_color(ws)
    ws.sheet_view.showGridLines = False

    hold = prop.hold_period_years
    noi_yr0 = calc_noi(prop.gross_rental_income, prop.vacancy_rate, prop.operating_expenses)
    annual_ds = calc_ds(prop.loan_amount, prop.interest_rate, prop.amortization_years)
    base_cap = calc_cap(noi_yr0, prop.purchase_price)
    exit_cap = prop.exit_cap_rate if prop.exit_cap_rate is not None else base_cap

    # Column widths
    _set_col_widths(ws, {"A": 4, "B": 32, "C": 16})
    for col_idx in range(4, 4 + hold + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    col_total = 3   # C
    col_yr0 = 4     # D

    # ---- Row 1: sheet title ----
    _write_section_header(ws, 1, 2, "Real Estate DCF Model (Annual)", span=3 + hold)

    # ---- Row 3: section label ----
    r3 = ws.cell(row=3, column=2, value="Annual Cash Flow (USD)")
    r3.font = _font(bold=True, size=11)

    # ---- Row 4: column headers ----
    ws.cell(row=4, column=2, value="").fill = _fill(TEAL_DARK)
    tc = ws.cell(row=4, column=col_total, value="Total")
    tc.fill = _fill(TEAL_DARK)
    tc.font = _font(bold=True, white=True)
    tc.alignment = _align("center")
    for yr in range(0, hold + 1):
        col = col_yr0 + yr
        hc = ws.cell(row=4, column=col, value=f"Year {yr}")
        hc.fill = _fill(TEAL_DARK)
        hc.font = _font(bold=True, white=True)
        hc.alignment = _align("center")

    # ---- Build per-year arrays ----
    occ_rate_vals:  list = [None]    # Year 0 = N/A
    gpi_vals:       list = [0.0]
    vac_loss_vals:  list = [0.0]
    egi_vals:       list = [0.0]
    other_inc_vals: list = [0.0]
    rev_vals:       list = [0.0]
    pm_fee_vals:    list = [0.0]
    opex_vals:      list = [0.0]
    noi_vals:       list = [0.0]
    capex_vals:     list = [0.0]
    ncf_vals:       list = [0.0]
    unlev_vals:     list = [0.0]
    ds_vals:        list = [0.0]
    btcf_vals:      list = [0.0]

    for yr in range(1, hold + 1):
        gpi = prop.gross_rental_income * ((1 + prop.rent_growth_annual) ** yr)
        vac_loss = gpi * prop.vacancy_rate
        egi = gpi - vac_loss
        opex = prop.operating_expenses * ((1 + prop.expense_growth_annual) ** yr)
        pm_fee = opex * 0.4
        noi = egi - opex
        capex = prop.capital_expenditures * ((1 + prop.expense_growth_annual) ** yr)
        ncf = noi - capex

        occ_rate_vals.append(1 - prop.vacancy_rate)
        gpi_vals.append(gpi)
        vac_loss_vals.append(-vac_loss)
        egi_vals.append(egi)
        other_inc_vals.append(0.0)
        rev_vals.append(egi)
        pm_fee_vals.append(-pm_fee)
        opex_vals.append(-opex)
        noi_vals.append(noi)
        capex_vals.append(-capex)
        ncf_vals.append(ncf)
        unlev_vals.append(ncf)
        ds_vals.append(-annual_ds)
        btcf_vals.append(ncf - annual_ds)

    # Exit
    noi_exit = noi_yr0 * ((1 + prop.rent_growth_annual) ** hold)
    exit_price = noi_exit / exit_cap
    disp_costs = exit_price * 0.02
    net_sale = exit_price - disp_costs
    remaining_debt = calc_rem(prop.loan_amount, prop.interest_rate, prop.amortization_years, hold)
    net_equity = net_sale - remaining_debt

    sale_vals    = [0.0] * (hold + 1)
    disp_vals    = [0.0] * (hold + 1)
    nsp_vals     = [0.0] * (hold + 1)
    debt_rep     = [0.0] * (hold + 1)
    nep_vals     = [0.0] * (hold + 1)
    sale_vals[hold]   = exit_price
    disp_vals[hold]   = -disp_costs
    nsp_vals[hold]    = net_sale
    debt_rep[hold]    = -remaining_debt
    nep_vals[hold]    = net_equity

    total_cf_row = [-prop.equity_invested]
    for yr in range(1, hold + 1):
        if yr < hold:
            total_cf_row.append(btcf_vals[yr])
        else:
            total_cf_row.append(btcf_vals[yr] + net_equity)

    def _sum_series(vals: list) -> float:
        return sum(v for v in vals if isinstance(v, (int, float)))

    def _write_cf_row(row: int, label: str, total_val, year_vals: list,
                      bold: bool = False, fmt: str = FMT_CURRENCY,
                      row_fill: Optional[str] = None,
                      indent: bool = False) -> None:
        lbl_text = ("  " if indent else "") + label
        lc = ws.cell(row=row, column=2, value=lbl_text)
        lc.font = _font(bold=bold)
        lc.alignment = _align("left")

        tv = ws.cell(row=row, column=col_total, value=total_val)
        tv.font = _font(bold=bold)
        tv.alignment = _align("right")
        if total_val is not None and isinstance(total_val, (int, float)):
            tv.number_format = fmt
        if row_fill:
            tv.fill = _fill(row_fill)

        for yr_idx in range(hold + 1):
            col = col_yr0 + yr_idx
            val = year_vals[yr_idx] if yr_idx < len(year_vals) else None
            vc = ws.cell(row=row, column=col, value=val)
            vc.font = _font(bold=bold)
            vc.alignment = _align("right")
            if val is not None and isinstance(val, (int, float)):
                vc.number_format = fmt
            if row_fill:
                vc.fill = _fill(row_fill)

    row = 6
    _write_cf_row(row, "Occupancy Rate", None, occ_rate_vals, fmt=FMT_PERCENT); row += 1
    _write_cf_row(row, "GPI (Gross Potential Income)", _sum_series(gpi_vals), gpi_vals); row += 1
    _write_cf_row(row, "  Vacancy Loss", _sum_series(vac_loss_vals), vac_loss_vals, indent=True); row += 1
    _write_cf_row(row, "EGI (Effective Gross Income)", _sum_series(egi_vals), egi_vals, bold=True,
                  row_fill=GRAY_CALC); row += 1
    _write_cf_row(row, "  Other Income", 0.0, other_inc_vals, indent=True); row += 1
    _write_cf_row(row, "Total Revenue", _sum_series(rev_vals), rev_vals, bold=True); row += 1

    row += 1  # blank
    _write_cf_row(row, "  PM Fee (est.)", _sum_series(pm_fee_vals), pm_fee_vals, indent=True); row += 1
    _write_cf_row(row, "  OPEX (Operating Expenses)", _sum_series(opex_vals), opex_vals, indent=True); row += 1
    _write_cf_row(row, "NOI", _sum_series(noi_vals), noi_vals, bold=True, row_fill=GRAY_CALC); row += 1
    _write_cf_row(row, "  CapEx", _sum_series(capex_vals), capex_vals, indent=True); row += 1
    _write_cf_row(row, "NCF", _sum_series(ncf_vals), ncf_vals, bold=True); row += 1
    _write_cf_row(row, "Unlevered CF", _sum_series(unlev_vals), unlev_vals); row += 1

    row += 1  # blank
    _write_cf_row(row, "  Debt Service", _sum_series(ds_vals), ds_vals, indent=True); row += 1
    _write_cf_row(row, "BTCF (Before-Tax Levered CF)", _sum_series(btcf_vals), btcf_vals,
                  bold=True); row += 1

    row += 1  # blank
    _write_cf_row(row, "Sale Price", None, sale_vals); row += 1
    _write_cf_row(row, "  Disposition Costs", None, disp_vals, indent=True); row += 1
    _write_cf_row(row, "Net Sale Proceeds", None, nsp_vals); row += 1
    _write_cf_row(row, "  Loan Balance Repayment", None, debt_rep, indent=True); row += 1
    _write_cf_row(row, "Net Equity Proceeds", None, nep_vals); row += 1

    row += 1  # blank
    _write_cf_row(row, "Total CF (to Equity)", None, total_cf_row, bold=True,
                  row_fill=YELLOW_KEY); row += 1

    # Highlight Year 0 in total CF row as equity outlay
    ws.cell(row=row - 1, column=col_yr0).fill = _fill(YELLOW_KEY)

    row += 1  # blank

    # Investment Metrics Summary sub-table
    _write_section_header(ws, row, 2, "Investment Metrics Summary", span=3)
    row += 1

    irr_val = metrics.get("irr")
    em_val = metrics.get("equity_multiple", 0.0)

    def _metric_row(r: int, label: str, value, fmt: Optional[str]) -> None:
        lc = ws.cell(row=r, column=2, value=label)
        lc.font = _font(bold=True)
        vc = ws.cell(row=r, column=col_total, value=value)
        vc.font = _font(bold=True)
        vc.fill = _fill(TEAL_MID)
        vc.alignment = _align("right")
        if fmt and value is not None and not isinstance(value, str):
            vc.number_format = fmt

    _metric_row(row, "Levered IRR (Pre-Tax)", irr_val, FMT_PERCENT); row += 1
    _metric_row(row, "Equity Multiple", em_val, FMT_MULTIPLE); row += 1
    _metric_row(row, "IRR (Levered, Annualized)", irr_val, FMT_PERCENT); row += 1


# ---------------------------------------------------------------------------
# Sheet 5: Sources & Uses
# ---------------------------------------------------------------------------

def _build_sources_uses(wb: Workbook, prop, metrics: dict) -> None:
    calc_noi, calc_cap, calc_ds, calc_rem = _calc_imports()

    ws = wb.create_sheet("Sources & Uses")
    _set_tab_color(ws)
    ws.sheet_view.showGridLines = False

    _set_col_widths(ws, {"A": 4, "B": 32, "C": 16, "D": 12, "E": 4, "F": 28, "G": 16, "H": 12})

    pp = prop.purchase_price

    # Acquisition cost estimates
    brokerage    = pp * 0.010
    reg_tax      = pp * 0.005
    re_acq_tax   = pp * 0.005
    stamp_duty   = pp * 0.002
    scrivener    = pp * 0.003
    appraisal    = pp * 0.002
    fire_ins     = pp * 0.001
    other_costs  = 0.0
    total_acq_costs = brokerage + reg_tax + re_acq_tax + stamp_duty + scrivener + appraisal + fire_ins + other_costs
    total_uses = pp + total_acq_costs

    total_debt = prop.loan_amount
    total_equity = prop.equity_invested
    total_sources = total_debt + total_equity
    diff = total_uses - total_sources

    # ---- Row 1-2: headers ----
    _write_section_header(ws, 1, 2, "Sources & Uses — Financing & Capital Structure", span=7)
    _write_section_header(ws, 2, 2, "Detailed financing terms and capital allocation", span=7)

    # ---- Row 4: dual-block column headers ----
    uses_headers = ["Uses (Capital Deployment)", "Amount", "% of Total"]
    src_headers  = ["Sources (Financing)", "Amount", "% of Total"]
    for ci, lbl in enumerate(uses_headers):
        c = ws.cell(row=4, column=2 + ci, value=lbl)
        c.fill = _fill(TEAL_DARK)
        c.font = _font(bold=True, white=True)
        c.alignment = _align("center")
    for ci, lbl in enumerate(src_headers):
        c = ws.cell(row=4, column=6 + ci, value=lbl)
        c.fill = _fill(TEAL_DARK)
        c.font = _font(bold=True, white=True)
        c.alignment = _align("center")

    def _su_row(row: int, label: str, amount: Optional[float],
                label_col: int = 2, amount_col: int = 3, pct_col: int = 4,
                total_base: float = total_uses,
                bold: bool = False, fill_clr: str = TEAL_LIGHT) -> None:
        lc = ws.cell(row=row, column=label_col, value=label)
        lc.font = _font(bold=bold)
        if amount is not None:
            ac = ws.cell(row=row, column=amount_col, value=amount)
            ac.font = _font(bold=bold)
            ac.fill = _fill(fill_clr)
            ac.alignment = _align("right")
            ac.number_format = FMT_CURRENCY
            if total_base and total_base > 0:
                pc = ws.cell(row=row, column=pct_col, value=amount / total_base)
                pc.font = _font(bold=bold)
                pc.fill = _fill(fill_clr)
                pc.alignment = _align("right")
                pc.number_format = FMT_PERCENT

    # ---- Uses block ----
    r = 5
    ws.cell(row=r, column=2, value="Acquisition Price").font = _font(bold=True)
    ws.cell(row=r, column=6, value="Debt (Borrowings)").font = _font(bold=True)
    r += 1

    _su_row(r, "  Property Price",              pp,                           fill_clr=TEAL_LIGHT)
    _su_row(r, "  Senior Loan",                 prop.loan_amount,             label_col=6, amount_col=7, pct_col=8, total_base=total_sources, fill_clr=TEAL_LIGHT)
    r += 1

    _su_row(r, "  (No Consumption Tax)",        0.0,                          fill_clr=TEAL_LIGHT)
    _su_row(r, "  Mezzanine",                   0.0,                          label_col=6, amount_col=7, pct_col=8, total_base=total_sources, fill_clr=TEAL_LIGHT)
    r += 1

    _su_row(r, "  Acquisition Price (total)",   pp,                           bold=True, fill_clr=GRAY_CALC)
    _su_row(r, "  Total Debt",                  total_debt,                   label_col=6, amount_col=7, pct_col=8, total_base=total_sources, bold=True, fill_clr=GRAY_CALC)
    r += 2  # blank + next section

    ws.cell(row=r, column=2, value="Acquisition Costs").font = _font(bold=True)
    ws.cell(row=r, column=6, value="Equity (Own Capital)").font = _font(bold=True)
    r += 1

    _su_row(r, "  Brokerage Fee (est. 1%)",     brokerage,    fill_clr=TEAL_LIGHT)
    _su_row(r, "  Equity Contribution",         total_equity, label_col=6, amount_col=7, pct_col=8, total_base=total_sources, fill_clr=TEAL_LIGHT)
    r += 1
    _su_row(r, "  Registration Tax (est. 0.5%)", reg_tax,     fill_clr=TEAL_LIGHT); r += 1
    _su_row(r, "  RE Acquisition Tax (est. 0.5%)", re_acq_tax, fill_clr=TEAL_LIGHT); r += 1
    _su_row(r, "  Stamp Duty (est.)",            stamp_duty,  fill_clr=TEAL_LIGHT); r += 1
    _su_row(r, "  Judicial Scrivener (est.)",    scrivener,   fill_clr=TEAL_LIGHT); r += 1
    _su_row(r, "  Appraisal & DD (est.)",        appraisal,   fill_clr=TEAL_LIGHT); r += 1
    _su_row(r, "  Fire Insurance (est.)",        fire_ins,    fill_clr=TEAL_LIGHT); r += 1
    _su_row(r, "  Other Costs",                  other_costs, fill_clr=TEAL_LIGHT); r += 1
    _su_row(r, "  Total Acquisition Costs",      total_acq_costs, bold=True, fill_clr=GRAY_CALC)
    _su_row(r, "  Total Sources",                total_sources, label_col=6, amount_col=7, pct_col=8, total_base=total_sources, bold=True, fill_clr=GRAY_CALC)
    r += 2

    ws.cell(row=r, column=2, value="Reserves").font = _font(bold=True)
    ws.cell(row=r, column=6, value="Check (Uses - Sources)").font = _font(bold=True)
    ck = ws.cell(row=r, column=7, value=diff)
    ck.number_format = FMT_CURRENCY
    ck.fill = _fill(GRAY_CALC)
    ck.alignment = _align("right")
    r += 1

    _su_row(r, "  Repair Reserve",   0.0, fill_clr=TEAL_LIGHT); r += 1
    _su_row(r, "  Working Capital",  0.0, fill_clr=TEAL_LIGHT); r += 1
    _su_row(r, "  Total Reserves",   0.0, bold=True, fill_clr=GRAY_CALC); r += 2

    _su_row(r, "Total Uses", total_uses, bold=True, fill_clr=GRAY_CALC); r += 3

    # ---- Loan Terms Detail ----
    annual_ds = calc_ds(prop.loan_amount, prop.interest_rate, prop.amortization_years)
    ltv_val = metrics.get("ltv", 0.0)

    ws.cell(row=r, column=2, value="Loan Terms Detail").font = _font(bold=True, size=11)
    r += 1

    loan_hdr_labels = ["Item", "Senior Loan", "Mezzanine Loan"]
    for ci, lbl in enumerate(loan_hdr_labels):
        c = ws.cell(row=r, column=2 + ci, value=lbl)
        c.fill = _fill(TEAL_DARK)
        c.font = _font(bold=True, white=True)
        c.alignment = _align("center")
    r += 1

    loan_rows = [
        ("Loan Amount",             prop.loan_amount,    0,      FMT_CURRENCY),
        ("LTV",                     ltv_val,             0,      FMT_PERCENT),
        ("Interest Type",           "Fixed",             "N/A",  None),
        ("Interest Rate",           prop.interest_rate,  0,      FMT_PERCENT),
        ("Loan Term",               f"{prop.loan_term_years} yrs", "N/A", None),
        ("Repayment Method",        "Fully Amortizing",  "N/A",  None),
        ("P&I Payment (annual)",    annual_ds,           0,      FMT_CURRENCY),
        ("Lender",                  "Commercial Bank (assumed)", "N/A", None),
        ("Collateral",              "1st Priority Mortgage",     "N/A", None),
    ]
    for label, senior_val, mezz_val, fmt in loan_rows:
        lc = ws.cell(row=r, column=2, value=label)
        lc.font = _font()

        sv = ws.cell(row=r, column=3, value=senior_val)
        sv.fill = _fill(TEAL_LIGHT)
        sv.alignment = _align("right")
        if fmt and isinstance(senior_val, (int, float)):
            sv.number_format = fmt

        mv = ws.cell(row=r, column=4, value=mezz_val)
        mv.fill = _fill(TEAL_LIGHT)
        mv.alignment = _align("right")
        if fmt and isinstance(mezz_val, (int, float)):
            mv.number_format = fmt
        r += 1


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def render_excel(memo_data: dict, output_path: str) -> None:
    """Generate Excel workbook matching the 不動産モデルTemplate.en.xlsx format.

    Sheets produced:
      1. Cover             — title page
      2. Input             — assumptions & parameters
      3. Property Summary  — asset overview, metrics, DCF, sensitivity
      4. CFs               — detailed annual levered cash flow model
      5. Sources & Uses    — financing & capital structure breakdown

    Args:
        memo_data: dict with keys 'property' (PropertyInput) and 'metrics' (dict).
        output_path: destination .xlsx file path.
    """
    prop = memo_data["property"]
    metrics = memo_data["metrics"]

    wb = Workbook()
    # Remove default blank sheet
    wb.remove(wb.active)

    _build_cover(wb, prop)
    _build_input(wb, prop)
    _build_property_summary(wb, prop, metrics)
    _build_cfs(wb, prop, metrics)
    _build_sources_uses(wb, prop, metrics)

    wb.save(output_path)
